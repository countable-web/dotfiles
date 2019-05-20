#!/bin/python

import httplib2
import os
import io

from apiclient import discovery
from apiclient.http import MediaIoBaseDownload

from oauth2client import client, file, tools
from oauth2client.file import Storage

import openpyxl
from openpyxl import Workbook

# import employees

try:
    import argparse
    flags = argparse.ArgumentParser(parents=[tools.argparser]).parse_args()
except ImportError:
    flags = None

# If modifying these scopes, delete your previously saved credentials
# at ~/.credentials/drive-python-quickstart.json
SCOPES = 'https://www.googleapis.com/auth/drive.readonly'
CLIENT_SECRET_FILE = 'client_id.json'
APPLICATION_NAME = 'Drive API Python Quickstart'


class SheetManager:
    EMPLOYEE_RATES = {}

    def __init__(self):
        credentials = self.get_google_drive_credentials()
        http = credentials.authorize(httplib2.Http())
        self.drive_service = discovery.build('drive', 'v3', http=http)

    def get_google_drive_credentials(self):
        """Gets valid user credentials from storage.

        If nothing has been stored, or if the stored credentials are invalid,
        the OAuth2 flow is completed to obtain the new credentials.

        Returns:
                        Credentials, the obtained credential.
        """
        home_dir = os.path.expanduser('~')
        credential_dir = os.path.join(home_dir, '.credentials')
        if not os.path.exists(credential_dir):
            os.makedirs(credential_dir)
        credential_path = os.path.join(
            credential_dir, 'client_id.json')
        # credential_dir, 'drive-python-quickstart.json')

        store = Storage(credential_path)
        credentials = store.get()
        if not credentials or credentials.invalid:
            flow = client.flow_from_clientsecrets(CLIENT_SECRET_FILE, SCOPES)
            flow.user_agent = APPLICATION_NAME
            if flags:
                credentials = tools.run_flow(flow, store, flags)
            else:  # Needed only for compatibility with Python 2.6
                credentials = tools.run(flow, store)
            print('Storing credentials to ' + credential_path)

        return credentials

    def get_pay(self):

        sheets_store = file.Storage('token.json')
        sheets_creds = sheets_store.get()
        if not sheets_creds or sheets_creds.invalid:
            flow = client.flow_from_clientsecrets('client_id.json', SCOPES)
            #flow = client.flow_from_clientsecrets('sheets_credentials.json', SCOPES)
            sheets_creds = tools.run_flow(flow, sheets_store)
        service = discovery.build(
            'sheets', 'v4', http=sheets_creds.authorize(httplib2.Http()))

        SPREADSHEET_ID = '1ibMFHmCZiHBRfIsAfL7npyxhK32mxkkcTuvLLdfM0_A'
      # Update RANGE_NAME when number of employees change
        RANGE_NAME = 'Rates!A3:B100'
        result = service.spreadsheets().values().get(spreadsheetId=SPREADSHEET_ID,
                                                     range=RANGE_NAME).execute()
        values = result.get('values', [])

        self.EMPLOYEE_RATES = {}

        if not values:
            print('No data found.')
        else:
            print('Getting Employee Rates')
            for row in values:
                SheetManager.EMPLOYEE_RATES[row[0]] = int(row[1])

            print(SheetManager.EMPLOYEE_RATES)
            return SheetManager.EMPLOYEE_RATES

    def search_timesheets(self, q):
        page_token = None
        while True:
            response = self.drive_service.files().list(q=q,
                                                       spaces='drive',
                                                       fields='nextPageToken, files(id, name, parents)',
                                                       pageToken=page_token).execute()
            for file in response.get('files', []):
                # Process change
                print('Found file: %s' % (file.get('name')))
                self.download_sheet(file)
            page_token = response.get('nextPageToken', None)
            if page_token is None:
                break

    def download_sheet(self, f):

        request = self.drive_service.files().export_media(
            fileId=f.get('id'),
            mimeType='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        fh = io.FileIO('tmp/{}.xlsx'.format(f.get('name')), 'w')
        downloader = MediaIoBaseDownload(fh, request)
        done = False
        while done is False:
            status, done = downloader.next_chunk()
            print("Download %d%%." % int(status.progress() * 100))

    def tally_hours(self, months=[
        'JAN', 'FEB', 'MAR',
        'APR', 'MAY', 'JUN',
        'JUL', 'AUG', 'SEP',
        'OCT', 'NOV', 'DEC'
    ]):
        employees_data = {}

        employee_files = []
        for filename in os.listdir("tmp/"):
            if filename.endswith(".xlsx"):
                employee_files.append(filename)

        output_book = Workbook()

        for filename in employee_files:
            print(filename)
            employee_name = filename.split("-")[0]
            output_sheet = output_book.create_sheet(employee_name, 0)

            if not filename.startswith('.'):

                wb = openpyxl.load_workbook("tmp/" + filename, data_only=True)
                total = 0
                for month_idx, month in enumerate(months):
                    try:
                        current_sheet = wb[month]
                    except KeyError:
                        continue

                    output_sheet.cell(month_idx + 2, 1).value = month

                    null_line = 0
                    projects = {}
                    start_row = 6
                    date_col = 2
                    project_col = 1
                    if current_sheet.cell(5, 5).value == 'Hours Worked':
                        hours_col = 5
                    elif current_sheet.cell(18, 6).value == 'Hours Worked':
                        hours_col = 6
                        start_row = 19
                        date_col = 1
                        project_col = 2
                    else:
                        hours_col = 6
                    for row in range(start_row, current_sheet.max_row):
                        cellValue = current_sheet.cell(
                            row, hours_col).value
                        project_name = current_sheet.cell(
                            row, project_col).value
                        the_date = current_sheet.cell(row, date_col).value

                        # if cellValue is None:
                        #     null_line += 1
                        #     if null_line == 2:
                        #         break
                        if not project_name:
                            if the_date:
                                print("**** no project name ****",
                                      the_date, cellValue)
                            continue
                        elif the_date:
                            # if project_name not in projects:
                            if project_name not in projects:
                                projects[project_name] = 0
                            null_line = 0
                            try:
                                projects[project_name] += float(
                                    cellValue or '0')
                            except:
                                print("*****", cellValue,
                                      'was a bad cell value in', month)
                    #print(row, 'rows')
                    row_idx = 1
                    for k, v in projects.items():
                        row_idx += 1
                        output_sheet.cell(1, row_idx).value = k
                        output_sheet.cell(month_idx + 2, row_idx).value = v

                    print(month)
        # employees_data[employee_name] = {
        #     'hours': total,
        #     'wage': SheetManager.EMPLOYEE_RATES[employee_name],
        #     'payment': total * SheetManager.EMPLOYEE_RATES[employee_name]
        # }

        # output_data = [['Employee', 'Hours', 'Wage', 'Payment']]
        # for employee in iter(employees_data):
        #     new_row = []
        #     new_row.append(employee)
        #     new_row.append(employees_data[employee]['hours'])
        #     new_row.append(employees_data[employee]['wage'])
        #     new_row.append(employees_data[employee]['payment'])
        #     output_data.append(new_row)
        # for row in output_data:
        #     rowIndex = output_data.index(row)
        #     for col in range(len(output_data[0])):
        #         output_sheet.cell(row=rowIndex + 1, column=col +
        #                             1).value = output_data[rowIndex][col]

        output_book.save(filename="PAYMENTS.xlsx")


def main():
    manager = SheetManager()
    #manager.search_timesheets("name contains '2018-timesheet'")
    manager.get_pay()
    manager.tally_hours()


if __name__ == '__main__':
    main()
